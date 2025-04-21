from flask import Flask, request, jsonify, send_from_directory, send_file
from bpp_aco_ms import load_boxes_from_json, Truck, Box, pack_multiple_trucks
from io import BytesIO
from datetime import datetime, timezone
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import sqlite3

app = Flask(__name__)

# Статические файлы сервера (HTML, JS). Начальный маршрут для приложения
@app.route('/')
def serve_index():
    return send_from_directory('static', 'index.html')

# Endpoint для получения стртовых данных по пакетам(коробкам)
@app.route('/api/boxes', methods=['GET'])
def get_boxes():
    boxes = load_boxes_from_json('boxes.json')
    boxes_data = [
        {
            'id': box.id,
            'length': box.length,
            'width': box.width,
            'height': box.height,
            'value': box.value,
            'weight': box.weight
        } for box in boxes
    ]
    return jsonify(boxes_data)

# Endpoint для получения стртовых данных по грузовикам
@app.route('/api/trucks', methods=['GET'])
def get_trucks():
    try:
        with open('trucks.json', 'r') as f:
            trucks_data = json.load(f)
        return jsonify(trucks_data)
    except Exception as e:
        return jsonify({'error': f'Ошибка загрузки грузовиков: {str(e)}'}), 500

# Endpoint для запуска упаковки
@app.route('/api/pack', methods=['POST'])
def pack_trucks():
    data = request.get_json()
    num_ants = int(data.get('num_ants', 50))
    iterations = int(data.get('iterations', 100))
    strategy = data.get('strategy', 'value_per_volume')
    alpha = float(data.get('alpha', 3.0))
    beta = float(data.get('beta', 2.0))
    evaporation_rate = float(data.get('evaporation_rate', 0.5))

    # Получаем данные грузовиков и пакетов из запроса
    boxes_data = data.get('boxes', [])
    trucks_data = data.get('trucks', [])

    # Преобразуем данные в объекты
    boxes = [
        Box(id=b['id'], length=b['length'], width=b['width'], height=b['height'],
            value=b['value'], weight=b['weight']) for b in boxes_data
    ]
    trucks = [
        Truck(length=t['length'], width=t['width'], height=t['height'], 
              max_weight=t['max_weight'], id=t['id']) for t in trucks_data
    ]

    # Упаковка
    used_trucks, unplaced_boxes = pack_multiple_trucks(
        trucks, boxes, num_ants, iterations, strategy, alpha, beta, evaporation_rate
    )

    result = []
    for truck in used_trucks:
        placed_boxes = [
            {
                'id': box.id,
                'length': box.length,
                'width': box.width,
                'height': box.height,
                'value': box.value,
                'weight': box.weight,
                'position': {'x': pos.x, 'y': pos.y, 'z': pos.z}
            } for box, pos in truck.placed_boxes
        ]
        result.append({
            'truck_id': truck.id,
            'length': truck.length,
            'width': truck.width,
            'height': truck.height,
            'max_weight': truck.max_weight,
            'placed_boxes': placed_boxes,
            'fitness': sum(b.length * b.width * b.height for b, _ in truck.placed_boxes) /
                       (truck.length * truck.width * truck.height),
            'total_weight': sum(b.weight for b, _ in truck.placed_boxes),
            'total_value': sum(b.value for b, _ in truck.placed_boxes)
        })

    unplaced = [{'id': b.id, 'length': b.length, 'width': b.width, 'height': b.height,
                 'value': b.value, 'weight': b.weight} for b in unplaced_boxes]

    parameters = {
        'num_ants': num_ants,
        'iterations': iterations,
        'strategy': strategy,
        'alpha': alpha,
        'beta': beta,
        'evaporation_rate': evaporation_rate
    }

    # Сохраняем в БД
    conn = sqlite3.connect('packing_results.db')
    cur = conn.cursor()
    timestamp = datetime.now(timezone.utc).isoformat()
    cur.execute('''
        INSERT INTO packing_results (timestamp, parameters, result_json)
        VALUES (?, ?, ?)
    ''', (timestamp, json.dumps(parameters), json.dumps({'trucks': result, 'unplaced_boxes': unplaced})))
    test_id = cur.lastrowid
    conn.commit()
    conn.close()

    return jsonify({'test_id': test_id, 'trucks': result, 'unplaced_boxes': unplaced})

# Endpoint для сохраннеия результата упаковки в Excel файл (экспериментально)
@app.route('/api/save_excel', methods=['POST'])
def save_excel():
    data = request.get_json()
    # извлекаем номер теста
    test_id = int(data.get('test_id'))
    
    if not test_id:
        return jsonify({'error': 'test_id обязателен'}), 400

    # Извлекаем результат из БД
    conn = sqlite3.connect('packing_results.db')
    cur = conn.cursor()
    cur.execute('SELECT parameters, result_json FROM packing_results WHERE id = ?', (test_id,))
    row = cur.fetchone()
    conn.close()

    if not row:
        return jsonify({'error': f'Нет данных с test_id = {test_id}'}), 404

    parameters = json.loads(row[0])
    result = json.loads(row[1])

    wb = Workbook()
    ws_params = wb.active
    ws_params.title = 'Параметры'

    # Пишем параметры алгоритма
    ws_params.append(['Параметр', 'Значение'])
    for key, value in parameters.items():
        ws_params.append([key, value])

    # Лист с грузовиками
    for truck in result['trucks']:
        ws = wb.create_sheet(title=f"Грузовик {truck['truck_id']}")
        ws.append(['Размеры', f"{truck['length']} x {truck['width']} x {truck['height']}"])
        ws.append(['Макс. вес', truck['max_weight']])
        ws.append(['Суммарный вес', truck['total_weight']])
        ws.append(['Суммарная ценность', truck['total_value']])
        ws.append(['Fitness (заполненность)', truck['fitness']])
        ws.append([])
        ws.append(['ID', 'Длина', 'Ширина', 'Высота', 'Вес', 'Ценность', 'X', 'Y', 'Z'])

        for box in truck['placed_boxes']:
            pos = box['position']
            ws.append([
                box['id'], box['length'], box['width'], box['height'],
                box['weight'], box['value'], pos['x'], pos['y'], pos['z']
            ])

    # Лист с неразмещёнными коробками
    ws_unplaced = wb.create_sheet(title='Неразмещённые')
    ws_unplaced.append(['ID', 'Длина', 'Ширина', 'Высота', 'Вес', 'Ценность'])
    for box in result['unplaced_boxes']:
        ws_unplaced.append([
            box['id'], box['length'], box['width'], box['height'],
            box['weight'], box['value']
        ])

    # Формируем файл
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Отправляем файл
    return send_file(
        output,
        as_attachment=True,
        download_name=f"packing_report_{test_id}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == '__main__':
    app.run(host="localhost", debug=True)